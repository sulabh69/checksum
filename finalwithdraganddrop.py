import tkinter as tk
from tkinterdnd2 import TkinterDnD, DND_FILES
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def browse_file(label, side):
    filename = filedialog.askopenfilename(title=f"Select the {side} Excel file", filetypes=[("Excel files", "*.xlsx")])
    if filename:
        label.config(text=filename)
    return filename

def drop_file(event, label):
    filename = event.data
    if filename.endswith('.xlsx'):
        label.config(text=filename)
    else:
        messagebox.showerror("Error", "Please drop an Excel file (.xlsx)")

def compare_sheets():
    file1 = left_label.cget("text")
    file2 = right_label.cget("text")

    if not file1 or not file2 or "Drag" in file1 or "Drag" in file2:
        messagebox.showerror("Error", "Please select or drop both files before comparing.")
        return

    try:
        df1 = pd.read_excel(file1)
        df2 = pd.read_excel(file2)
        comparison = df1.compare(df2, keep_shape=True, keep_equal=True)

        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if output_file:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df1.to_excel(writer, sheet_name='Comparison', index=False)

            wb = load_workbook(output_file)
            ws = wb['Comparison']
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

            for row in range(2, len(df1) + 2):
                for col in range(1, len(df1.columns) + 1):
                    cell_value1 = ws.cell(row=row, column=col).value
                    cell_value2 = df2.iloc[row-2, col-1] if row-2 < len(df2) and col-1 < len(df2.columns) else None
                    if pd.notna(cell_value1) and pd.notna(cell_value2) and cell_value1 != cell_value2:
                        ws.cell(row=row, column=col).fill = red_fill

            wb.save(output_file)

            messagebox.showinfo("Success", f"Comparison complete. Results saved to {output_file}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

app = TkinterDnD.Tk()
app.title("Excel Comparator")

# Left Section (File to Compare)
left_frame = tk.Frame(app, padx=20, pady=20, bg="lightblue")
left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

left_label = tk.Label(left_frame, text="Drag and drop or click to upload the file to compare", bg="lightgray", width=40, height=10, relief=tk.RAISED, borderwidth=2)
left_label.pack(fill=tk.BOTH, expand=True)
left_label.drop_target_register(DND_FILES)
left_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, left_label))

left_button = tk.Button(left_frame, text="Upload File", command=lambda: browse_file(left_label, "first"))
left_button.pack(pady=10)

# Right Section (Comparison File)
right_frame = tk.Frame(app, padx=20, pady=20, bg="lightgreen")
right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

right_label = tk.Label(right_frame, text="Drag and drop or click to upload the comparison file", bg="lightgray", width=40, height=10, relief=tk.RAISED, borderwidth=2)
right_label.pack(fill=tk.BOTH, expand=True)
right_label.drop_target_register(DND_FILES)
right_label.dnd_bind('<<Drop>>', lambda event: drop_file(event, right_label))

right_button = tk.Button(right_frame, text="Upload File", command=lambda: browse_file(right_label, "second"))
right_button.pack(pady=10)

# Compare Button
compare_button = tk.Button(app, text="Compare Excel Sheets", command=compare_sheets, pady=10)
compare_button.pack(pady=20)

app.mainloop()
